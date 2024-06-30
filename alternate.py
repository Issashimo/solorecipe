import openpyxl


def alternate(ingredients):

    write_wb = openpyxl.load_workbook("data/recipe.xlsx")
    write_ws = write_wb["Sheet2"]

    alternate = []
    texts = []
    for i in range(write_ws.max_row):
        if write_ws.cell(row=i + 1, column=1).value not in ingredients:
            j = 2
            judge = True
            while write_ws.cell(row=i + 1, column=j).value != None:
                if write_ws.cell(row=i + 1, column=j).value in ingredients:
                    pass
                else:
                    judge = False
                    break
                j += 1
            if judge == True:
                alternate.append(write_ws.cell(row=i + 1, column=1).value)
                text = f"{write_ws.cell(row=i + 1, column=1).value},{write_ws.cell(row=i + 1, column=6).value}\n"
                texts.append(text)

    with open("data/alternate.txt", mode="w") as f:
        f.writelines(texts)
    return alternate
