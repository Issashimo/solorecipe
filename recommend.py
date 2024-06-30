def recommender(ingredients):
    import openpyxl
    import alternate
    import random

    alternates = alternate.alternate(ingredients)

    ingredients_alt = ingredients
    for i in alternates:
        ingredients_alt.append(i)

    write_wb = openpyxl.load_workbook("data/Rakuten_recipe.xlsx")
    write_ws = write_wb["Sheet1"]
    scores = []

    for i in range(write_ws.max_row):
        score_recipe = []
        j = 0
        score = 0
        while True:
            if write_ws.cell(row=i + 1, column=7 + j).value in ingredients_alt:
                score += 1
                j += 1
            elif write_ws.cell(row=i + 1, column=7 + j).value == None:
                break
            else:
                pass
                j += 1

        try:
            score_recipe.append(score / j + 0.01 * j + random.uniform(-0.001, 0.001))

        except:
            score_recipe.append(0)

        score_recipe.append(write_ws.cell(row=i + 1, column=2).value)
        score_recipe.append(write_ws.cell(row=i + 1, column=3).value)
        score_recipe.append(write_ws.cell(row=i + 1, column=4).value)
        scores.append(score_recipe)

    scores_sorted = sorted(scores, reverse=True)

    result = []
    for i in range(6):
        result.append(scores_sorted[i])

    return result
