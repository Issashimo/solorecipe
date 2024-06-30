import requests
import json
import time
import pandas as pd
import openpyxl
from pprint import pprint


write_wb = openpyxl.load_workbook("data/Rakuten_recipe.xlsx")
write_ws = write_wb["Sheet1"]

df_recipe = pd.DataFrame(
    columns=[
        "recipeId",
        "recipeTitle",
        "foodImageUrl",
        "recipeMaterial",
        "recipeCost",
        "recipeIndication",
    ]
)

import API

df_keyword = API.ID()
count = 0

for i in df_keyword:
    time.sleep(
        3
    )  # 連続でアクセスすると先方のサーバに負荷がかかるので少し待つのがマナー

    url = (
        "https://app.rakuten.co.jp/services/api/Recipe/CategoryRanking/20170426?applicationId=1080534497721355158&categoryId="
        + str(i)
    )
    res = requests.get(url)

    json_data = json.loads(res.text)
    recipes = json_data["result"]

    for recipe in recipes:
        df_recipe = df_recipe._append(
            {
                "recipeId": recipe["recipeId"],
                "recipeTitle": recipe["recipeTitle"],
                "foodImageUrl": recipe["foodImageUrl"],
                "recipeMaterial": recipe["recipeMaterial"],
                "recipeCost": recipe["recipeCost"],
                "recipeIndication": recipe["recipeIndication"],
            },
            ignore_index=True,
        )
    count += 1

    print(abs(100 * count / len(df_keyword)))

for i in range(4 * len(df_keyword)):
    try:
        write_ws.cell(row=write_ws.max_row + 1, column=1, value=i)
        write_ws.cell(row=write_ws.max_row, column=2, value=df_recipe.at[i, "recipeId"])
        write_ws.cell(
            row=write_ws.max_row, column=3, value=df_recipe.at[i, "recipeTitle"]
        )
        write_ws.cell(
            row=write_ws.max_row, column=4, value=df_recipe.at[i, "foodImageUrl"]
        )
        write_ws.cell(
            row=write_ws.max_row, column=5, value=df_recipe.at[i, "recipeCost"]
        )
        write_ws.cell(
            row=write_ws.max_row, column=6, value=df_recipe.at[i, "recipeIndication"]
        )
        count_Material = 0
        for j in df_recipe.at[i, "recipeMaterial"]:
            write_ws.cell(row=write_ws.max_row, column=7 + count_Material, value=j)
            count_Material += 1

    except:
        pass

write_wb.save("data/Rakuten_recipe.xlsx")
