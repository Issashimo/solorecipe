from bs4 import BeautifulSoup
import requests
import openpyxl
import link
import time


def page(number):
    links = link.link("https://cookpad.com/category/10" + str(number))

    write_wb = openpyxl.load_workbook("recipe.xlsx")
    write_ws = write_wb["Sheet1"]

    for i in links:
        try:
            url = "https://cookpad.com" + i

            response = requests.get(url)
            html_content = response.text

            soup = BeautifulSoup(html_content, "html.parser")

            title = soup.find(class_="recipe-title")

            ingredient = soup.find_all(class_="ingredient_name")

            write_ws.cell(row=write_ws.max_row + 1, column=1, value=title.text)
            write_ws.cell(row=write_ws.max_row, column=2, value=url)
            count = 0
            for i in ingredient:
                write_ws.cell(row=write_ws.max_row, column=3 + count, value=i.text)
                count += 1
            
        except:
            pass

        time.sleep(1)

    write_wb.save("recipe.xlsx")


if __name__ == "__main__":
    link()
